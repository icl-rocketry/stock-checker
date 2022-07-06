import pygsheets
from mouser.api import MouserPartSearchRequest
from datetime import datetime
from openpyxl import load_workbook

API_KEYS_FILE = 'C:\stockchecker\mouser_api_keys.yaml'

#Google Sheets
#gc = pygsheets.authorize(service_file='auth.json')
#sh = gc.open('Component Stock Checking')
#wks = sh[0]
#read = wks.get_as_df()

#Excel
stocksheet = load_workbook(filename="stock.xlsx")
sheet = stocksheet.worksheets[0]
datalen = sheet.max_row-1

stend = ' In Stock'

stock = []
for i in range(0,datalen):
    args = []
    partno = str(sheet['A'+str(2+i)].value) #excel
    qneeded = int(sheet['D'+str(2+i)].value)
    #print(read.iloc[i][0]) #google sheets
    print(partno)
    request = MouserPartSearchRequest('partnumber', API_KEYS_FILE, *args)
    if request.url:
            # Run request
            search = request.part_search(partno)
            if search:
                # Print result
                try:
                    respraw = request.get_clean_response()
                    for j in range (0,len(respraw["PriceBreaks"])-1):
                        if j < len(respraw["PriceBreaks"])-1:
                            qlow = int(respraw["PriceBreaks"][j]["Quantity"])
                            qhigh = int((respraw["PriceBreaks"][j+1]["Quantity"]))
                            if qneeded >= qlow and qneeded < qhigh:
                                price = float(respraw["PriceBreaks"][j]["Price"].removeprefix('£'))
                        else:
                            price = float(respraw["PriceBreaks"][j]["Price"].removeprefix('£'))
                            
                    #stock.append([respraw["Availability"]])
                    sheet["B"+str(2+i)] = respraw["ProductDetailUrl"]
                    sheet["C"+str(2+i)] = price
                    sheet["F"+str(2+i)] = int(respraw["Availability"].removesuffix(stend)) #excel
                    
                except Exception:
                    pass

now = datetime.now()
current_time = now.strftime("%H:%M:%S on %d/%m/%Y")

#Google Sheets
#wks.update_values((2,6),stock) 
#wks.update_values('L2', [[current_time]])

#Excel
sheet["L2"] = current_time
stocksheet.save(filename="stock.xlsx")